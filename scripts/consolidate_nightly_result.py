# SPDX-FileCopyrightText: (c) 2025 Tenstorrent AI ULC
#
# SPDX-License-Identifier: Apache-2.0
from datetime import date, timedelta
import requests
import os
import io
import re
import json
import zipfile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Constants
GITHUB_TOKEN = "your access token"  # Replace with your github token
REPO_OWNER = "tenstorrent"
REPO_NAME = "tt-forge-fe"

# GitHub API endpoint to get workflow runs
nightly_url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/actions/runs?per_page=100"

# Headers for authentication
headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}

# Intialization
all_out_dict = {}
failure_reason_lines = []
latest_10_nightly_id = []
latest_10_run_id = []
total_count = 0
fail_count = 0
xfail_count = 0
skip_count = 0
pass_count = 0
runner_count = 4
log_path_list = []
latest_job_name_list = []
header_name = ["JOB_NAME"]


def unzip_files(output_file):
    """
    Function is used to unzip log files
    ----------
    Parameters
    ----------
    output_file : File which is need to be unzip
    """
    file_count = 0
    for item in os.listdir(output_file):
        if item.endswith(".zip"):
            file_count = file_count + 1
            new_file_name = "pytest_" + str(file_count) + ".log"
            extracted_file_path = os.path.join(output_file, new_file_name)
            zip_file_path = os.path.join(output_file, item)

            with zipfile.ZipFile(zip_file_path, "r") as zip_ref:
                for file_info in zip_ref.infolist():
                    with zip_ref.open(file_info.filename) as file:
                        with open(extracted_file_path, "wb") as f_out:
                            f_out.write(file.read())
            os.remove(zip_file_path)


def download_ci_logs(owner, repo, run_id, output_file):
    """
    Function is used to download the log files
    ----------
    Parameters
    ----------
    owner       : owner of the repo
    repo        : Repo name
    run_id      : run id which is used to download the log
    output_file : path where to store the output logs
    """
    artifact_url = f"https://api.github.com/repos/{owner}/{repo}/actions/runs/{run_id}/artifacts"
    artifact_url_response = requests.get(artifact_url, headers=headers, stream=True)
    if artifact_url_response.status_code == 200:
        try:
            response_str = artifact_url_response.content.decode("utf-8")
            artifacts = json.loads(response_str).get("artifacts", [])

            if artifacts:
                for artifact in artifacts:
                    if "test-log-n150" in str(artifact["name"]):
                        download_url = artifact["archive_download_url"]
                        download_response = requests.get(download_url, headers=headers)
                        if download_response.status_code == 200:
                            artifact_file_name = output_file + "/" + str(artifact["name"]) + ".zip"
                            with open(artifact_file_name, "wb") as f:
                                f.write(download_response.content)

                        else:
                            print(f"Error downloading artifact: {download_response.status_code}")

                unzip_files(output_file)

            else:
                print("No artifacts found.")

        except requests.exceptions.JSONDecodeError as e:
            print(f"JSONDecodeError: {str(e)}")
            print("The response is not valid JSON.")
    else:
        print(f"Failed to download logs. HTTP Status: {artifact_url_response.status_code}")


def fetch_nightly_ids(url):
    """
    Function is used to get latest 10 nightly id's
    ----------
    Parameters
    ----------
    url : url which is used to fetch nightly data's
    """
    id_count = 0

    while url:
        # Fetch workflow runs
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()
            runs = data.get("workflow_runs", [])

            if runs:

                for run in runs:

                    if run["name"] == "On nightly":
                        if run["head_branch"] == "main" and run["event"] == "schedule":
                            id_count += 1
                            nightly_date = str((date.today() - timedelta(id_count)).strftime("%m-%d")).replace("-", "/")
                            nightly_id_date = str(run["run_number"]) + "-" + nightly_date
                            latest_10_nightly_id.append(nightly_id_date)
                            latest_10_run_id.append(str(run["id"]))
                        if id_count > 9:
                            break

                if id_count > 9:
                    break

            if "next" in response.links:
                # Set the 'next' page URL for the next request
                url = response.links["next"]["url"]
            else:
                # No more pages, break the loop
                url = None

        else:
            print(f"Error: {response.status_code}, {response.text}")

    print("LATEST 10 NIGHTLY ID's :", latest_10_nightly_id)


def write_excelfile(content, ws):
    """
    Function is used to write excel sheet
    ----------
    Parameters
    ----------
    content : Data dictionary
    ws      : Need to write data on this worksheet
    """

    if ws.title == "Failure_Reason":
        print("Writing Failure_Reason worksheet ...")
        ws.append(["JOB_NAME", "STATUS"])
        for jb_name, info in content.items():
            ws.append([jb_name] + [info])

    elif ws.title == "All_Runner_Results":
        print("Writing All runner worksheet ...")
        ws.append(header_name)

        for jb_name, info in content.items():
            ws.append([jb_name] + info)
    align_excel(ws)


def align_excel(content_sheet):
    """
    Aligning excel sheet
    ----------
    Parameter
    ----------
    content_sheet   : Sheet needs to be aligned
    """
    working_sheet = content_sheet

    # Adding Border
    side = Side(style="thin", color="000000")
    thin_border = Border(left=side, right=side, top=side, bottom=side)
    for row in working_sheet.iter_rows(
        min_row=1, max_row=working_sheet.max_row, min_col=1, max_col=working_sheet.max_column
    ):
        for cell in row:
            cell.border = thin_border

    # Header Alignment
    blue_fill = PatternFill(start_color="d4e2f8", end_color="d4e2f8", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center")
    for col in range(1, working_sheet.max_column + 1):
        working_sheet.cell(row=1, column=col).fill = blue_fill
        working_sheet.cell(row=1, column=col).font = Font(bold=True)
        working_sheet.cell(row=1, column=col).alignment = center_aligned

    # Calculating and Assigning column width
    for col_num in range(1, working_sheet.max_column + 1):
        max_length = 0

        for row_num in range(1, working_sheet.max_row + 1):
            cell = working_sheet.cell(row=row_num, column=col_num)

            # Calculate the length of the cell's content (if it's not None)
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        column_letter = get_column_letter(col_num)
        working_sheet.column_dimensions[column_letter].width = max_length + 2

    if working_sheet.title == "Failure_Reason":
        for row_index in range(2, working_sheet.max_row + 1):
            working_sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="left", vertical="center")

            # Calculating and Assigning row height
            line_count = 0
            cell = working_sheet.cell(row=row_index, column=2)
            list_value = str(cell.value).split("\n")
            for line in list_value:
                if line:
                    line_count = line_count + 1

            column_letter = get_column_letter(col_num)
            working_sheet.row_dimensions[row_index].height = 20 * line_count

    elif working_sheet.title == "All_Runner_Results":
        choices = ["PASS", "XPASS", "FAIL", "XFAIL", "SKIP"]
        dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"', showDropDown=True)
        working_sheet.add_data_validation(dv)

        # Fixing color for DataValidation choices
        pass_fill = PatternFill(start_color="afd8bc", end_color="afd8bc", fill_type="solid")  # light green
        xpass_fill = PatternFill(start_color="1c8348", end_color="1c8348", fill_type="solid")  #  dark Green
        fail_fill = PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid")  # light Red
        xfail_fill = PatternFill(start_color="f3c4cf", end_color="f3c4cf", fill_type="solid")  # dark red
        skip_fill = PatternFill(start_color="ffe599", end_color="ffe599", fill_type="solid")  # light yello
        for row_index in range(2, working_sheet.max_row + 1):
            working_sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="left", vertical="center")
            for col_index in range(2, working_sheet.max_column + 1):
                cell = working_sheet.cell(row=row_index, column=col_index)
                dv.add(cell)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.value == "PASS":
                    cell.fill = pass_fill
                elif cell.value == "XPASS":
                    cell.fill = xpass_fill
                elif cell.value == "FAIL":
                    cell.fill = fail_fill
                elif cell.value == "XFAIL":
                    cell.fill = xfail_fill
                elif cell.value == "SKIP":
                    cell.fill = skip_fill


def add_testcase(test_name, value, latest_result):
    """
    Adding data in dictionary
    ----------
    Parameters
    ----------
    test_name       : passing line in which testcase name is presented
    value           : status of the testcase ('PASS','XPASS','FAIL','XFAIL','SKIP')
    latest_result   : result is latest one or not (True or False)
    """
    extract_testname = (test_name.split("py::"))[1]
    if latest_result == True:
        latest_job_name_list.append(extract_testname)
        all_out_dict[extract_testname] = [value]
        test_log_status = "ended"
    else:
        if extract_testname in latest_job_name_list:
            all_out_dict[extract_testname].append(value)
        test_log_status = "ended"

    return test_log_status


def get_failure_details(failure_reason_lines):
    """
    Getting Failure Reasons in detail
    ----------
    Parameter
    ----------
    failure_reason_lines    : List of failure lines
    """
    out_dict = {}
    last_test_line = ""
    extra_error_lines = 3
    extract_error = []
    index_cnt = 0
    test_logs = False
    for line in failure_reason_lines:
        if "_ test_" in line:
            test_logs = True
            if line in out_dict:
                print("Reading same test skipped")
            else:
                last_test_line = line

        if line.startswith("E  ") and (test_logs == True) or ("[XPASS(strict)]" in line):
            for i in range(-extra_error_lines, (extra_error_lines + 1)):
                extract_error.append(failure_reason_lines[index_cnt - i])

            key = last_test_line.strip("_ ")
            if "[XPASS(strict)]" in line:
                out_dict[key] = "[XPASS(strict)]"
            else:
                out_dict[key] = "\n".join(extract_error) + "\n"
            extract_error = []
            test_logs = False

        index_cnt = index_cnt + 1

    write_excelfile(out_dict, new_ws_fr)


if __name__ == "__main__":
    fetch_nightly_ids(nightly_url)
    latest_id = latest_10_nightly_id[0]

    # Creating Worksheet
    new_wb = Workbook()
    initial_ws = new_wb.active
    new_wb.remove(initial_ws)
    new_ws = new_wb.create_sheet(title="All_Runner_Results")
    new_ws_fr = new_wb.create_sheet(title="Failure_Reason")

    # Creating Output Directory
    output_dir_date = (date.today() - timedelta(1)).isoformat()
    output_first_dir = "Nightly_results/" + output_dir_date
    os.makedirs(output_first_dir, exist_ok=True)
    run_count = 0

    for id in latest_10_nightly_id:

        if latest_id == id:
            latest_result = True
        else:
            latest_result = False

        log_path_list = []
        output_second_dir = "Nightly_results/" + output_dir_date + "/" + str(id).split("-")[0]
        os.makedirs(output_second_dir, exist_ok=True)
        run_id = latest_10_run_id[run_count]
        download_ci_logs(REPO_OWNER, REPO_NAME, run_id, output_second_dir)
        run_count = run_count + 1

        dir = os.listdir(output_second_dir)
        if len(dir) != 0:
            header_name.append(str(id))
            for i in range(1, (runner_count + 1)):
                runner_path = output_second_dir + "/pytest_" + str(i) + ".log"
                log_path_list.append(runner_path)

            for log_file in log_path_list:
                with open(log_file, "r") as log:
                    lines = log.readlines()

                append_lines = False
                iteration = False
                test_log_status = "ended"

                for line in lines:
                    if "==== short test summary info ====" in line:
                        break

                    if "warnings summary" in line:
                        continue

                    if "plugins: " in line:
                        iteration = True
                    elif iteration == False:
                        continue

                    if ("forge/test/" in line) and (iteration == True):
                        test_case_name = line.strip("\n").split(" ")
                        test_log_status = "started"

                    if ("XFAIL" in line) and (test_log_status == "started"):
                        test_case_name[0] = test_case_name[0].strip("XFAIL")
                        test_log_status = add_testcase(test_case_name[0], "XFAIL", latest_result)
                        if latest_result == True:
                            total_count = total_count + 1
                            xfail_count = xfail_count + 1

                    if ("PASSED" in line) and (test_log_status == "started"):
                        test_log_status = add_testcase(test_case_name[0], "PASS", latest_result)
                        if latest_result == True:
                            total_count = total_count + 1
                            pass_count = pass_count + 1

                    if ("FAILED" in line) and (test_log_status == "started"):
                        test_log_status = add_testcase(test_case_name[0], "FAIL", latest_result)
                        if latest_result == True:
                            total_count = total_count + 1
                            fail_count = fail_count + 1

                    if (" SKIPPED" in line) and (test_log_status == "started"):
                        test_log_status = add_testcase(test_case_name[0], "SKIP", latest_result)
                        if latest_result == True:
                            total_count = total_count + 1
                            skip_count = skip_count + 1

                    if latest_result == True:
                        if "=== FAILURES ====" in line:
                            append_lines = True

                        if append_lines:
                            failure_reason_lines.append(line.strip("\n"))

                        if "warnings summary" in line:
                            break
            cmd3 = f"rm -rf {output_second_dir}"
        else:
            print("-------Empty directory for run id:", id, "-------")
            cmd3 = f"rmdir {output_second_dir}"

        os.system(cmd3)

    # Need to print latest result alone
    print(
        "\n==============================" + "\033[1m",
        "Short Test Summary",
        "\033[0m" + "=============================",
    )
    print("TOTAL COUNT :", total_count)
    print("PASS COUNT  :", pass_count)
    print("FAIL COUNT  :", fail_count)
    print("XFAIL COUNT :", xfail_count)
    print("SKIP COUNT  :", skip_count)
    print("================================================================================")
    write_excelfile(all_out_dict, new_ws)
    get_failure_details(failure_reason_lines)

    # Saving the workbook
    new_wb.save(output_first_dir + "/Nightly_workbook.xlsx")
    print("\nNightly_workbook.xlsx was created\n")
