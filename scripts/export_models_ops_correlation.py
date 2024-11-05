# SPDX-FileCopyrightText: (c) 2024 Tenstorrent AI ULC
# SPDX-License-Identifier: Apache-2.0
import sys
import os
import subprocess
import argparse
from loguru import logger
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font

test_to_skip = []


def collect_all_pytests(root_dir_path):

    logger.info(f"Collecting all pytests in {root_dir_path}")

    try:
        res = subprocess.check_output(["pytest", root_dir_path, "--setup-plan"], stderr=subprocess.STDOUT).decode(
            "utf-8"
        )
    except subprocess.CalledProcessError as e:
        output = e.output.decode("utf-8")
        logger.error(f"[Error!] output = {output}")
        return []

    test_list = []
    lines = res.split("\n")
    for line in lines:
        if "warnings summary" in line or "slowest durations" in line:
            break

        if line and line.startswith("        " + root_dir_path) and "::" in line and "training" not in line:
            line = line.strip()
            line = line.split(" (fixtures used:")[0] if " (fixtures used:" in line else line
            if "Grayskull" not in line and "Wormhole_B0" not in line:
                test_list.append(line)

    return test_list


def run_and_export_models_unique_ops_configuration(
    compilation_depth, pytest_directory_path, export_unique_op_config_file_type
):

    # Collect all the pytests in the directory path by the user
    test_list = collect_all_pytests(pytest_directory_path)
    assert test_list != [], "No tests found!"

    # Run all the pytests upto compilation_depth provided by the user, which will dump all the ops configuration as CSV File
    # in current working directory or path which is set using the FORGE_EXPORT_UNIQUE_OP_CONFIG_DIR_PATH flag.
    for test in test_list:

        if test in test_to_skip:
            continue

        logger.info(f"Running the test: {test} upto {compilation_depth.lower()} compilation depth...")

        try:
            result = subprocess.run(
                [
                    "pytest",
                    test,
                    "-vss",
                ],
                capture_output=True,
                text=True,
                check=True,
                env=dict(
                    os.environ,
                    FORGE_COMPILE_DEPTH=compilation_depth,
                    FORGE_EXTRACT_UNIQUE_OP_CONFIG_AT=compilation_depth.upper(),
                    FORGE_EXPORT_UNIQUE_OP_CONFIG_FILE_TYPE=export_unique_op_config_file_type,
                    FORGE_DISABLE_REPORTIFY_DUMP="1",
                    FORGE_EXPORT_UNIQUE_OP_CONFIG_DIR_PATH=os.getenv(
                        "FORGE_EXPORT_UNIQUE_OP_CONFIG_DIR_PATH", os.getcwd()
                    ),
                ),
            )
            if result.returncode != 0:
                logger.error(f"Error while running the pytest:\n stdout: {result.stdout}\n stderr: {result.stderr}")
            else:
                logger.info(f"Successfully ran the test")

        except subprocess.CalledProcessError as e:
            logger.error(f"Error while running the pytest:\n {e.output}")


def create_sheet_from_exported_models_unique_op_configuration(
    compilation_depth, output_directory_path, output_file_name, export_unique_op_config_file_type
):

    logger.info("Creating a sheet for model variants unique op configurations...")

    # Get the exported models unique op config directory path
    exported_unique_op_config_dir_path = os.getenv("FORGE_EXPORT_UNIQUE_OP_CONFIG_DIR_PATH", os.getcwd())
    if "OpConfigs" not in exported_unique_op_config_dir_path:
        exported_unique_op_config_dir_path = os.path.join(exported_unique_op_config_dir_path, "OpConfigs")

    unique_ops_dict = dict()
    models_list = os.listdir(exported_unique_op_config_dir_path)
    for model_name in models_list:
        model_path = os.path.join(exported_unique_op_config_dir_path, model_name)
        dumped_op_configuration_files = os.listdir(model_path)
        for dumped_op_configuration_file in dumped_op_configuration_files:

            if not dumped_op_configuration_file.endswith(export_unique_op_config_file_type):
                continue

            if (
                dumped_op_configuration_file.replace("." + export_unique_op_config_file_type, "").lower()
                == compilation_depth.lower()
            ):
                dumped_op_configuration_file_path = os.path.join(model_path, dumped_op_configuration_file)

                if export_unique_op_config_file_type == "csv":
                    op_config_df = pd.read_csv(
                        dumped_op_configuration_file_path,
                        sep=os.getenv("FORGE_EXPORT_UNIQUE_OP_CONFIG_CSV_DELIMITER", "/"),
                        header=0,
                        usecols=["OpName", "Shape", "Attributes"],
                    )

                elif export_unique_op_config_file_type == "xlsx":
                    op_config_df = pd.read_excel(
                        dumped_op_configuration_file_path,
                        header=0,
                        usecols=["OpName", "Shape", "Attributes"],
                    )

                for index, row in op_config_df.iterrows():
                    op_name = str(row.OpName)
                    shape = str(row.Shape)
                    attr = str(row.Attributes)

                    shape = shape.replace("{", "[").replace("}", "]")
                    if shape.endswith(",]"):
                        shape = shape[:-2] + "]"
                    if shape.endswith(", ]"):
                        shape = shape[:-3] + "]"

                    if op_name in unique_ops_dict.keys():
                        if shape not in unique_ops_dict[op_name].keys():
                            unique_ops_dict[op_name][shape] = {attr: 1}
                        else:
                            if attr not in unique_ops_dict[op_name][shape].keys():
                                unique_ops_dict[op_name][shape][attr] = 1
                            else:
                                unique_ops_dict[op_name][shape][attr] = unique_ops_dict[op_name][shape][attr] + 1

                    else:
                        unique_ops_dict[op_name] = {shape: {attr: 1}}

    # Convert Unique Op configuration to list of row for xlsx sheet
    data = list()
    for op_name, shape_attrs in unique_ops_dict.items():
        for shape, attrs in shape_attrs.items():
            for attr, num_ops in attrs.items():
                if attr == "nan":
                    attr = ""
                data.append([op_name, shape, attr, num_ops])

    max_column_width = max(len(str(item)) for row in data for item in row)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = output_file_name

    headers = ["OpName", "Shape", "Attributes", "NumOps"]
    sheet.append(headers)

    for row in data:
        sheet.append(row)

    blue_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center")
    side = Side(style="thin", color="000000")
    thin_border = Border(left=side, right=side, top=side, bottom=side)

    # Fill Header with blue color
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=col).fill = blue_fill
        sheet.cell(row=1, column=col).font = Font(bold=True)

    # Make thin border and center align the text for every cell
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).alignment = center_aligned

    # Set column width for cells
    column_offset = 2
    for col in sheet.columns:
        column = col[0].column_letter
        sheet.column_dimensions[column].width = max_column_width + column_offset

    # Check if the output directory exists otherwise create the directory
    if not os.path.exists(output_directory_path):
        os.makedirs(output_directory_path)

    workbook.save(os.path.join(output_directory_path, output_file_name + ".xlsx"))

    logger.info(
        f"Saved models unique op configuration to {os.path.join(output_directory_path,output_file_name+'.xlsx')}"
    )


def create_model_variants_and_ops_cross_correlation_sheet(
    compilation_depth, output_directory_path, output_file_name, export_unique_op_config_file_type
):

    logger.info("Creating a sheet for cross correlation between model variants and unique op configuration...")

    # Get the exported models unique op config directory path
    exported_unique_op_config_dir_path = os.getenv("FORGE_EXPORT_UNIQUE_OP_CONFIG_DIR_PATH", os.getcwd())
    if "OpConfigs" not in exported_unique_op_config_dir_path:
        exported_unique_op_config_dir_path = os.path.join(exported_unique_op_config_dir_path, "OpConfigs")

    # Create python dictionary which contains key as model names and values as list of ops names
    max_model_name_len = 0
    models_ops_dict = dict()
    models_list = os.listdir(exported_unique_op_config_dir_path)

    for model_name in models_list:

        # Calculate model name length for setting the column size
        if len(str(model_name)) > max_model_name_len:
            max_model_name_len = len(str(model_name))

        model_path = os.path.join(exported_unique_op_config_dir_path, model_name)

        dumped_op_configuration_files = os.listdir(model_path)
        for dumped_op_configuration_file in dumped_op_configuration_files:

            if not dumped_op_configuration_file.endswith(export_unique_op_config_file_type):
                continue

            if (
                dumped_op_configuration_file.replace("." + export_unique_op_config_file_type, "").lower()
                == compilation_depth.lower()
            ):

                dumped_op_configuration_file_path = os.path.join(model_path, dumped_op_configuration_file)

                if export_unique_op_config_file_type == "csv":
                    op_config_df = pd.read_csv(
                        dumped_op_configuration_file_path,
                        sep=os.getenv("FORGE_EXPORT_UNIQUE_OP_CONFIG_CSV_DELIMITER", "/"),
                        header=0,
                        usecols=["OpName", "Shape", "Attributes"],
                    )
                elif export_unique_op_config_file_type == "xlsx":
                    op_config_df = pd.read_excel(
                        dumped_op_configuration_file_path,
                        header=0,
                        usecols=["OpName", "Shape", "Attributes"],
                    )

                if model_name not in models_ops_dict.keys():
                    models_ops_dict[model_name] = op_config_df.OpName.tolist()
                else:
                    models_ops_dict[model_name].extend(op_config_df.OpName.tolist())

    all_models_ops_list = []
    for model, ops in models_ops_dict.items():
        all_models_ops_list.append(pd.DataFrame({"Models": model, "Ops": ops}))

    # Combine all into a single dataframe on index axis(i.e 0)
    combined_models_ops_df = pd.concat(all_models_ops_list)

    # Count occurrences of each operation across all models
    op_count = combined_models_ops_df["Ops"].value_counts().rename("OpCount")

    # Pivot to get Ops as rows and Models as columns
    pivot_df = pd.pivot_table(
        combined_models_ops_df,
        index="Ops",
        columns="Models",
        aggfunc="size",
        fill_value=0,
    )

    # Replace the counts with 'Y' for operations present, and 'N' for absent
    pivot_df = pivot_df.applymap(lambda x: "Y" if x > 0 else "N")

    # Add the operation counts as a new column
    pivot_df["OpCount"] = op_count

    # Sort rows by number of models supporting each operation
    pivot_df["model_ops_count"] = pivot_df.apply(lambda row: (row == "Y").sum(), axis=1)
    pivot_df = pivot_df.sort_values(by="model_ops_count", ascending=False).drop("model_ops_count", axis=1)

    # Sort columns by model names in alphabetical order
    column_names = list(pivot_df.columns)
    column_names.remove("OpCount")
    column_names.sort()
    column_names.append("OpCount")
    pivot_df = pivot_df[column_names]

    # Check if the output directory exists otherwise create the directory
    if not os.path.exists(output_directory_path):
        os.makedirs(output_directory_path)

    # Create a new Excel writer object
    writer = pd.ExcelWriter(
        os.path.join(output_directory_path, output_file_name + ".xlsx"),
        engine="openpyxl",
    )

    # Write the dataframe to the Excel file
    pivot_df.to_excel(writer, sheet_name=output_file_name)

    # Access the workbook and sheet
    workbook = writer.book
    ws = workbook.active
    sheet = writer.sheets[output_file_name]

    # Define your styles for fill the cell with particular color, center align the text in cell
    pastel_gray_fill = PatternFill(start_color="CED2BA", end_color="CED2BA", fill_type="solid")
    red_fill = PatternFill(start_color="D30000", end_color="D30000", fill_type="solid")
    blue_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
    slightly_desaturated_violet_fill = PatternFill(start_color="8D6FD1", end_color="8D6FD1", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center")
    side = Side(style="thin", color="000000")
    thin_border = Border(left=side, right=side, top=side, bottom=side)

    # Fill first column(i.e Ops) with blue color
    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=1).fill = blue_fill
        sheet.cell(row=row, column=1).border = thin_border
        sheet.cell(row=row, column=1).alignment = center_aligned

    # Fill first row(i.e Models) with blue color
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=col).fill = blue_fill
        sheet.cell(row=1, column=col).border = thin_border
        sheet.cell(row=1, column=col).alignment = center_aligned

    # Fill last column(i.e OpCount) with red color
    for rol in range(2, sheet.max_row + 1):
        sheet.cell(row=rol, column=sheet.max_column).fill = slightly_desaturated_violet_fill

    # Set column width for cells
    column_offset = 2
    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = max_model_name_len + column_offset

    # Loop over the cells in the sheet and apply color based on value
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_aligned
            if cell.value == "Y":
                cell.fill = pastel_gray_fill
            elif cell.value == "N":
                cell.fill = red_fill

    # Save the Excel file with formatting
    writer.save()
    writer.close()

    logger.info(
        f"Saved cross correlation details between model variants vs unique op configs to {os.path.join(output_directory_path,output_file_name+'.xlsx')}"
    )


if __name__ == "__main__":

    parser = argparse.ArgumentParser(
        description="Gather cross correlation details between model variants vs unique op configs and export model variants unique configuration to xlsx file"
    )
    parser.add_argument(
        "-c",
        "--compile_depth",
        choices=[
            "GENERATE_INITIAL_GRAPH",
            "POST_INITIAL_GRAPH_PASS",
            "OPTIMIZED_GRAPH",
            "AUTOGRAD",
            "POST_AUTOGRAD_PASS",
            "PRE_LOWERING_PASS",
        ],
        required=True,
        help="Choose the compilation depth for extracting ops configuration for the models present in pytest_directory_path.",
    )
    parser.add_argument(
        "-i",
        "--pytest_directory_path",
        required=True,
        help="Specify the directory path containing models to test.",
    )
    parser.add_argument(
        "--cross_correlation_output_file_name",
        default="Models_Ops_cross_correlation_data",
        required=False,
        help="Specify the output xlsx file name for saving the cross correation data between model variants and unique ops.",
    )
    parser.add_argument(
        "--models_unique_op_configs_output_file_name",
        default="Models_Unique_Op_Configurations",
        required=False,
        help="Specify the output xlsx file name for saving the Models unique op configurations.",
    )
    parser.add_argument(
        "-o",
        "--output_directory_path",
        default=os.getcwd(),
        required=False,
        help="Specify the output directory path for saving the xlsx file.",
    )
    parser.add_argument(
        "--export_unique_op_config_file_type",
        choices=["csv", "xlsx"],
        default="csv",
        required=False,
        help="Specify the export unique op configuration file type",
    )
    args = parser.parse_args()

    compilation_depth = args.compile_depth
    pytest_directory_path = args.pytest_directory_path
    cross_correlation_output_file_name = args.cross_correlation_output_file_name
    output_directory_path = args.output_directory_path
    models_unique_op_configs_output_file_name = args.models_unique_op_configs_output_file_name
    export_unique_op_config_file_type = args.export_unique_op_config_file_type

    run_and_export_models_unique_ops_configuration(
        compilation_depth, pytest_directory_path, export_unique_op_config_file_type
    )
    create_model_variants_and_ops_cross_correlation_sheet(
        compilation_depth, output_directory_path, cross_correlation_output_file_name, export_unique_op_config_file_type
    )
    create_sheet_from_exported_models_unique_op_configuration(
        compilation_depth,
        output_directory_path,
        models_unique_op_configs_output_file_name,
        export_unique_op_config_file_type,
    )
