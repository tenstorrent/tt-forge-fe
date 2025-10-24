# SPDX-FileCopyrightText: (c) 2025 Tenstorrent AI ULC
#
# SPDX-License-Identifier: Apache-2.0

import argparse
import json
from loguru import logger
import os
import sys
import tempfile
import xml.etree.ElementTree as ET
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font

# Default durations file used when none provided
DURATION_FILE = ".test_durations"


def create_excel_file(title: str, headers: List[str], rows: List[List[Any]], output_file_path: str) -> None:
    """
    Create a simple Excel file with headers and rows and basic styling.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]  # Excel sheet title length limit

    # Add headers
    sheet.append(headers)

    # Append rows
    for row in rows:
        # convert all cells to string for consistent sizing
        sheet.append([str(x) if x is not None else "" for x in row])

    # Styling
    blue_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side = Side(style="thin", color="000000")
    thin_border = Border(left=side, right=side, top=side, bottom=side)
    header_font = Font(bold=True)

    # Header style
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = blue_fill
        cell.font = header_font
        cell.alignment = center_aligned
        cell.border = thin_border

    # Row style
    for r in range(2, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r, column=c)
            cell.alignment = center_aligned
            cell.border = thin_border

    # Auto column widths (based on max length per column)
    col_max_lengths: List[int] = [len(h) for h in headers]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for idx, val in enumerate(row):
            length = len(str(val)) if val is not None else 0
            if idx < len(col_max_lengths):
                col_max_lengths[idx] = max(col_max_lengths[idx], length)
            else:
                col_max_lengths.append(length)

    # Apply widths (add a small padding)
    for i, width in enumerate(col_max_lengths, start=1):
        column_letter = sheet.cell(row=1, column=i).column_letter
        sheet.column_dimensions[column_letter].width = max(8, width + 2)

    # Save workbook
    workbook.save(output_file_path)
    logger.info("Saved Excel report to %s", output_file_path)


def load_json_file(path: str) -> Dict[str, Any]:
    """
    Load a JSON file and return the parsed object.
    Raises exception on error.
    """
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def atomic_write_json(path: str, data: Dict[str, Any]) -> None:
    """
    Atomically write JSON `data` to `path` using a temp file + os.replace.
    """
    ddir = os.path.dirname(os.path.abspath(path)) or "."
    os.makedirs(ddir, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=ddir)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, sort_keys=True)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except Exception as exc:
                logger.debug(f"Failed to remove temp file {tmp}: {exc}")


def load_existing_durations(path: str) -> Dict[str, Any]:
    """
    Load existging durations from `path` or return empty dict if missing or corrupted.
    """
    if not os.path.exists(path):
        logger.debug(f"Durations file {path} does not exist, returning empty dict")
        return {}
    try:
        return load_json_file(path)
    except Exception as exc:
        logger.warning(f"Failed to parse existing durations file {path}: {exc} — treating as empty")
        return {}


def extract_test_case_info(xml_file: str) -> Dict[str, float]:
    """
    Extract test case names and execution times from a JUnit XML report.

    Returns:
        dict mapping 'path.py::testname' -> duration_seconds (float)
    """
    test_cases_info: Dict[str, float] = {}
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()

        for testsuite in root.findall("testsuite"):
            # Iterate over all <testcase> elements within the current <testsuite>
            for testcase in testsuite.findall("testcase"):
                try:
                    path = testcase.get("classname").replace(".", "/")
                    name = testcase.get("name")
                    test_cases_info[f"{path}.py::{name}"] = float(testcase.get("time", 0))
                except ValueError:
                    print(f"Warning: Non-numeric time value encountered in {xml_file} for test case '{name}'")

    except ET.ParseError as e:
        logger.error(f"Error parsing XML file {xml_file}: {e}")
    except Exception as e:
        logger.error(f"Error reading file {xml_file}: {e}")

    return test_cases_info


def process_directory(directory: str) -> Dict[str, float]:
    """
    Walk directory and extract test durations from all .xml files that look like JUnit reports.
    """
    all_test_cases: Dict[str, float] = {}
    directory = os.path.abspath(directory)
    if not os.path.exists(directory):
        logger.error(f"Provided directory does not exist: {directory}")
        return {}

    for subdir, _, files in os.walk(directory):
        for fname in files:
            if fname.lower().endswith(".xml"):
                xml_file_path = os.path.join(subdir, fname)
                info = extract_test_case_info(xml_file_path)
                # merge (keeping maximum duration when same key appears)
                for k, v in info.items():
                    if k in all_test_cases:
                        all_test_cases[k] = max(all_test_cases[k], v)
                    else:
                        all_test_cases[k] = v
    logger.info(f"Processed directory {directory} — found {len(all_test_cases)} unique test cases")
    return all_test_cases


def update_test_duration(
    existing: Dict[str, float], collected: Dict[str, float], offset: Optional[float]
) -> Tuple[Dict[str, float], List[List[Any]]]:
    """
    Compare collected durations with existing durations and produce:
      - result dict (merged durations, using collected values as authoritative)
      - csv_data rows: [test_case, status(New/Updated), current_duration, previous_duration_or_empty]

    offset: if provided (>=0), a change smaller than offset is considered not significant.
    """
    csv_rows: List[List[Any]] = []
    result = dict(existing)  # start from existing

    # Add/Update from collected
    for test_case, new_duration in collected.items():
        prev_duration = existing.get(test_case)
        if prev_duration is None:
            # New test case
            result[test_case] = new_duration
            csv_rows.append([test_case, "NEW", new_duration, ""])
        else:
            # Already present: check if change is significant
            if offset is None:
                significant = new_duration != prev_duration
            else:
                try:
                    significant = abs(new_duration - float(prev_duration)) >= float(offset)
                except Exception:
                    significant = new_duration != prev_duration
            if significant:
                csv_rows.append([test_case, "UPDATED", new_duration, prev_duration])
                result[test_case] = new_duration
            # else: no change -> keep existing

    # Optionally, you might want to detect removed tests (present in existing but not in collected).
    # We'll not treat removals as updates. If you want removals listed, you can add them here.

    logger.info(f"Update results: {len(csv_rows)} updated/new rows")
    return result, csv_rows


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Collect test durations from JUnit XMLs and update duration and create XLSX report."
    )
    parser.add_argument(
        "--junit_xml_report_dir_path", required=True, type=str, help="Directory containing JUnit XML files."
    )
    parser.add_argument(
        "--out_file_path",
        required=False,
        default=DURATION_FILE,
        type=str,
        help="Output JSON file path for saving test durations.",
    )
    parser.add_argument(
        "--override_test_durations",
        action="store_true",
        help="If set, update existing durations only where collected differs (respecting offset). Otherwise, overwrite with collected durations.",
    )
    parser.add_argument(
        "--offset",
        required=False,
        type=float,
        default=None,
        help="Minimum difference required to consider a duration as changed (seconds).",
    )
    parser.add_argument(
        "--xlsx_report_file_path",
        required=False,
        type=str,
        help="If provided, an XLSX report will be created at this path.",
    )
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)

    logger.info("Starting test duration extraction")
    new_test_durations = process_directory(args.junit_xml_report_dir_path)
    if not new_test_durations:
        logger.error(f"No test durations found in directory: {args.junit_xml_report_dir_path}")
        return 2

    existing_test_durations = load_existing_durations(args.out_file_path)

    if args.override_test_durations:
        logger.info(f"Merging changes into existing durations using offset={args.offset}")
        merged, csv_rows = update_test_duration(existing_test_durations, new_test_durations, args.offset)
    else:
        # Overwrite mode: new durations replace existing completely
        logger.info("Overwrite mode: replacing existing durations with collected durations")
        merged = dict(new_test_durations)
        csv_rows = [[k, "NEW", v, ""] for k, v in merged.items()]

    # Write merged durations atomically
    try:
        atomic_write_json(args.out_file_path, merged)
        logger.info(f"Wrote durations JSON to {args.out_file_path} (entries={len(merged)})")
    except Exception as exc:
        logger.error(f"Failed to write durations JSON to {args.out_file_path}: {exc}")
        return 2

    # Optionally produce Excel
    if args.xlsx_report_file_path:
        headers = ["Test Case", "Status", "Current Test Duration (s)", "Previous Test Duration (s)"]
        create_excel_file(
            title="Test Duration Report", headers=headers, rows=csv_rows, output_file_path=args.xlsx_report_file_path
        )

    # Also print a small console summary
    logger.info("Summary:")
    logger.info(f"  Collected tests: {len(new_test_durations)}")
    logger.info(f"  Existing tests : {len(existing_test_durations)}")
    logger.info(f"  Merged tests   : {len(merged)}")
    logger.info(f"  New/Updated rows: {len(csv_rows)}")
    if args.xlsx_report_file_path:
        logger.info(f"  Excel report   : {args.xlsx_report_file_path}")

    return 0


if __name__ == "__main__":
    exit_code = main(sys.argv[1:])
    sys.exit(exit_code)
