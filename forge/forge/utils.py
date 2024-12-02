# SPDX-FileCopyrightText: © 2024 Tenstorrent AI ULC

# SPDX-License-Identifier: Apache-2.0

from typing import Dict, List, Optional

import filelock
import functools
import json
import torch
import math
import numpy as np
import hashlib
import getpass
import os
import shutil
import sys
import subprocess
import dataclasses_json
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font

from .forgeglobal import TILE_DIM

TILE_WIDTH = TILE_DIM
TILE_HEIGHT = TILE_DIM


def align_up(v, a):
    v -= 1
    return v - (v % a) + a


def align_up_tile(v):
    v -= 1
    return v - (v % TILE_WIDTH) + TILE_WIDTH


def round_up_div(n, d):
    return (n + d - 1) // d


def clamp(a, lower=None, upper=None):
    if lower is not None and a < lower:
        return lower
    if upper is not None and a > upper:
        return upper
    return a


def calculate_output_dimensions(original_x, original_y, stride, padding):
    import math

    return math.ceil(original_x / stride), math.ceil(original_y / stride)


# def generate_conv2d_inputs(input_channels, act_r_dim, act_c_dim, r0_rows_per_iteration=0):
#    padded_inp_channels = math.ceil(input_channels / TILE_HEIGHT) * TILE_HEIGHT
#    padded_r_dim = math.ceil(act_r_dim / TILE_HEIGHT) * TILE_HEIGHT
#    padded_c_dim = math.ceil(act_c_dim / TILE_WIDTH) * TILE_WIDTH
#
#    activations = np.random.uniform(low=0, high=0.1, size=(1, input_channels, act_r_dim, act_c_dim)).astype("float32")
#
#    # Transpose for forge, leaving the original activations for pytorch expected_result
#    npad = ( (0, 0), (0, padded_inp_channels - input_channels), (0, padded_r_dim - act_r_dim), (0, padded_c_dim - act_c_dim) )
#    act_padded = np.pad(activations, pad_width=npad, mode='constant', constant_values=0)
#    act_transposed = np.transpose(act_padded, (0, 2, 1, 3))
#    act_transposed = np.transpose(act_transposed, (0, 1, 3, 2))
#    act_transposed = np.copy(act_transposed, order='C') # settle the order before sending it to C
#    activation_rows = get_conv_tensor_row_count(act_r_dim, act_c_dim, r0_rows_per_iteration)
#    tt_act_transposed = flatten_for_conv(act_transposed, act_c_dim, act_r_dim, output_rows=activation_rows, shape_only=True)
#    tt_act_transposed_with_values = flatten_for_conv(act_transposed, act_c_dim, act_r_dim, output_rows=activation_rows, shape_only=False) # did not want to reimplement for python
#    act_transposed = tt_tensor_wrapper_to_numpy(tt_act_transposed_with_values)
#
#    return torch.tensor(activations), torch.tensor(act_transposed)
#
#
# def transform_forge_conv2d_output(result, act_c_dim, act_r_dim, stride, channel_size=None):
#    # Re-transpose back
#    act_c_dim = math.ceil(act_c_dim / stride)
#    act_r_dim = math.ceil(act_r_dim / stride)
#    result_as_tt_tensor_wrapper = tt_tensor_to_tt_tensor_wrapper(numpy_to_tt_tensor(result.numpy(), TensorType.Activation, False))
#    unflattened_result_as_tt_tensor_wrapper = unflatten_from_conv(result_as_tt_tensor_wrapper, act_c_dim, act_r_dim, False)
#    result = tt_tensor_wrapper_to_numpy(unflattened_result_as_tt_tensor_wrapper) # unflatted into 3d
#    result = np.transpose(result, (0, 1, 3, 2))
#    result = np.transpose(result, (0, 2, 1, 3))
#
#    if channel_size == None:
#        channel_size = result.shape[1]
#    result = result[:, :channel_size, :act_r_dim, :act_c_dim] # unpad
#    return result
#
# def extract_reduce_outputs(result, dim, keepdims):
#    dim_param_to_int = {
#        forge.Dim.R : (-2,),
#        forge.Dim.C : (-1,),
#        forge.Dim.Z : (-3,),
#        forge.Dim.RC : (-2, -1),
#        forge.Dim.ZR : (-3, -2),
#    }
#
#    operate_dim = dim_param_to_int[dim]
#    pick_index = torch.tensor([0,])
#
#    for i in operate_dim:
#        result = torch.index_select(result, i, pick_index)
#
#    if not keepdims:
#        for i in operate_dim:
#            result = torch.squeeze(result, i)
#
#    return result
#


def as_json(t):
    return dataclasses_json.config(encoder=t.to_json, decoder=t.from_json)


def dict_as_json(t):
    def to_json(d):
        return {k: t.to_json(v) for k, v in d.items()}

    def from_json(d):
        return {k: t.from_json(v) for k, v in d.items()}

    return dataclasses_json.config(encoder=to_json, decoder=from_json)


def list_as_json(t):
    if t is tuple:
        to_json = list
        from_json = tuple
    else:

        def to_json(d):
            return [t.to_json(v) for v in d]

        def from_json(d):
            return [t.from_json(v) for v in d]

    return dataclasses_json.config(encoder=to_json, decoder=from_json)


def optional_as_json(t):
    def to_json(d):
        return None if d is None else t.to_json(d)

    def from_json(d):
        return None if d is None else t.from_json(d)

    return dataclasses_json.config(encoder=to_json, decoder=from_json)


def get_padded_tensors(parameters):
    """Forge expects activation/parameter tensors to be 4-dimensions R/C-dim being 32-aligned"""

    updated_tensors = {}
    for parameter_name, parameter_tensor in parameters.items():
        while parameter_tensor.dim() < 4:
            parameter_tensor = torch.unsqueeze(parameter_tensor, 0)

        updated_tensors[parameter_name] = parameter_tensor
        """
        updated_tensors[parameter_name] = torch.nn.functional.pad(
            parameter_tensor,
            pad=(
                0,
                align_up_tile(parameter_tensor.shape[-1]) - parameter_tensor.shape[-1],
                0,
                align_up_tile(parameter_tensor.shape[-2]) - parameter_tensor.shape[-2],
            ),
        )
        """
    return updated_tensors


def get_forge_parameters_from_state_dict(state_dict: Dict[str, torch.Tensor]):
    from forge.parameter import Parameter

    forge_parameters = {}
    torch_parameters = get_padded_tensors(state_dict)
    for parameter_name, parameter_tensor in torch_parameters.items():
        forge_parameters[parameter_name] = Parameter(
            parameter_tensor,
            requires_grad=parameter_tensor.requires_grad,
        )
    return forge_parameters


def detach_tensors(tensors: List[torch.Tensor], fix_non_contiguos: bool = False) -> List[torch.Tensor]:
    """
    Detach tensors, and set requires_grad again if needed. Optionally clone non-contiguous tensors.
    """
    detached_tensors = [t.detach() for t in tensors]
    for t, old_t in zip(detached_tensors, tensors):
        t.requires_grad = old_t.requires_grad or old_t.grad_fn is not None

    if fix_non_contiguos:
        detached_tensors = [t if t.is_contiguous() else t.contiguous() for t in detached_tensors]
        assert all([t.is_contiguous() for t in detached_tensors])

    return detached_tensors


def get_forge_git_hash() -> Optional[str]:
    try:
        git_hash = (
            subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], stderr=subprocess.STDOUT)
            .decode("utf-8")
            .strip()
        )
        if git_hash.isalnum():
            return git_hash
        else:
            return None
    except:
        return None


def get_forgebackend_git_hash() -> Optional[str]:
    try:
        git_hash = (
            subprocess.check_output(
                ["git", "rev-parse", "--short", "HEAD:third_party/forgebackend"], stderr=subprocess.STDOUT
            )
            .decode("utf-8")
            .strip()
        )
        if git_hash.isalnum():
            return git_hash
        else:
            return None
    except:
        return None


def forgebackend_path() -> str:
    if "FORGE_HOME" in os.environ:
        return os.environ["FORGE_HOME"]

    if os.path.exists(os.getcwd() + "/third_party/forgebackend"):
        # must be in forge root
        return "third_party/forgebackend/"
    else:
        return ""


def resolve_device_descriptor_path(device_yaml_override: str) -> str:
    if os.path.isfile(device_yaml_override):
        return device_yaml_override
    elif os.path.isfile(forgebackend_path() + f"device/{device_yaml_override}"):
        return forgebackend_path() + f"device/{device_yaml_override}"
    else:
        raise FileNotFoundError(f"Device descriptor file not found: {device_yaml_override}")


def get_forge_compile_and_runtime_configs() -> Dict[str, str]:
    """
    Capture compile-time and runtime environment variables used to compile and run on the device.
    Eventually we want to separate out compile-time and runtime environment variables but we don't
    currently have a good way to do that yet.

    The current filter/capture is just a filter for 'FORGE_*' and 'TT_BACKEND_*'
    """
    compile_and_runtime_env_vars = {
        config: value for config, value in os.environ.items() if config.startswith(("FORGE_", "TT_BACKEND_"))
    }
    return compile_and_runtime_env_vars


def write_forge_envs_configs(dst_dir: str) -> None:
    with open(os.path.join(dst_dir, "compile_and_runtime_config.json"), "w") as json_file:
        json.dump(get_forge_compile_and_runtime_configs(), json_file, indent=4)


def get_tmp_dir() -> str:
    return os.path.join(os.environ.get("TMPDIR", "/tmp"), getpass.getuser())


def get_output_build_dir() -> str:
    user_defined_path = os.environ.get("FORGE_BUILD_DIR", None)
    output_build_directory = user_defined_path or get_tmp_dir()
    return output_build_directory


def get_lock_file_path(directory: str) -> str:
    parent_directory = os.path.abspath(os.path.join(directory, os.pardir))
    lock_file_name = f"{os.path.basename(directory)}.lock"
    return os.path.join(get_tmp_dir(), lock_file_name)


def file_lock_directory(func):
    @functools.wraps(func)
    def wrapper(test_output_directory, *args, **kwargs):
        lock_file_path = get_lock_file_path(test_output_directory)
        with filelock.FileLock(lock_file_path):
            return func(test_output_directory, *args, **kwargs)

    return wrapper


@file_lock_directory
def clear_test_output_build_directory(directory: str) -> None:
    try:
        if os.path.exists(directory):
            shutil.rmtree(directory)
    except Exception as e:
        logger.error(f"Failed to clear {directory}. Reason: {e}")


@file_lock_directory
def create_test_output_build_directory(directory: str) -> None:
    try:
        os.makedirs(directory, exist_ok=True)
    except Exception as e:
        logger.error(f"Failed to create test output build dir{directory}. Reason: {e}")


@file_lock_directory
def make_test_output_directory(test_output_directory: str) -> None:
    os.makedirs(test_output_directory, exist_ok=True)


def get_current_pytest():
    if "PYTEST_CURRENT_TEST" in os.environ:
        return os.environ.get("PYTEST_CURRENT_TEST", "").replace(" (call)", "")
    else:
        import sys

        return " ".join(sys.argv)


def generate_hash(org_str):
    return hashlib.sha256(org_str.encode()).hexdigest()[:12]


def resolve_output_build_directory(*, directory_prefix: str = None) -> str:
    """
    Return the path to the temp directory where the test output build artifacts will be dumped.
    Order of path resolution:
        1. Use user-defined path if set: FORGE_BUILD_DIR
        2. Default to TMPDIR environment variable
        3. If TMPDIR environment variable is unset, default to /tmp/
    """
    output_build_directory = get_output_build_dir()
    os.makedirs(output_build_directory, exist_ok=True)

    forge_env_options = get_forge_compile_and_runtime_configs()
    forge_env_options_hash = generate_hash(str(forge_env_options))

    test_permutation = forge_env_options_hash + get_current_pytest()
    hashed_suffix = generate_hash(test_permutation)

    test_output_directory_name = "_".join(filter(None, [directory_prefix, hashed_suffix]))
    test_output_directory = os.path.join(output_build_directory, test_output_directory_name)

    make_test_output_directory(test_output_directory)

    return test_output_directory


def create_excel_file(title: str, headers: List[str], rows: List[List[str]], output_file_path: str):

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)

    for row in rows:
        sheet.append(row)

    blue_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center")
    side = Side(style="thin", color="000000")
    thin_border = Border(left=side, right=side, top=side, bottom=side)

    # Fill Header with blue color
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=col).fill = blue_fill
        sheet.cell(row=1, column=col).font = Font(bold=True)

    # Make thin border and center align the text for every cell
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).alignment = center_aligned

    # Calculate width for each columns
    columns_width = {}
    for header in headers:
        columns_width[header] = len(header)
    for row in rows:
        for header, item in zip(headers, row):
            columns_width[header] = max(columns_width[header], len(str(item)))

    # Set column width for cells
    column_offset = 2
    for col, col_width in zip(sheet.columns, columns_width.values()):
        column = col[0].column_letter
        sheet.column_dimensions[column].width = col_width + column_offset

    workbook.save(output_file_path)


def create_xlsx_file_from_unique_op_config(
    unique_op_shapes_attrs, graph_name, stage, export_unique_op_config_file_path
):

    # Convert Unique Op configuration to list of row for xlsx sheet
    unique_op_config_data = []
    for op_name, shapes_attrs in unique_op_shapes_attrs.items():
        for shape, attrs in shapes_attrs:
            if len(attrs) != 0:
                for attr in attrs:
                    unique_op_config_data.append([str(op_name), str(shape), str(attr)])
            else:
                unique_op_config_data.append([str(op_name), str(shape), ""])

    sheet_title = graph_name + "_" + stage
    headers = ["OpName", "Shape", "Attributes"]

    create_excel_file(
        title=sheet_title,
        headers=headers,
        rows=unique_op_config_data,
        output_file_path=export_unique_op_config_file_path,
    )

    return True
